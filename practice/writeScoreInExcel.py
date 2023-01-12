"""
    假设某学校所有课程每学期允许多次考试，学生可随时参加考试，
    系统自动将每次成绩添加到 Excel 文件（包含 3 列：姓名，课程，成绩）中，
    现期末要求统计所有学生每门课程的最高成绩。
    编写程序，模拟生成若干同学的成绩并写入 Excel 文件，其中学生姓名和课程名称均可重复，
    也就是允许出现同一门课程的多次成绩，最后统计所有学生每门课程的最高成绩，
    并写入新的 Excel 文件。（建议考虑使用 openpyxl 并运用字典结构解决问题。）
"""
import random
from openpyxl import Workbook

# 课程表
subjects = ['高等数学', '线性代数', 'Java', 'Python', '运筹学']
# 获取学生花名册
with open(r'practice\names.txt', 'r', encoding='utf8') as fp:
    names = []
    for name in fp.readlines():
        names.append(name.strip())
    studentsNum = len(names)  # 学生人数
# 假设每门课程各有 100 次考试机会，且每人都至少参加一次考试
# 生成随机数据
scores = {}
for subject in subjects:
    for time in range(100):
        if time < studentsNum:
            stu = time
            # 这个 if 条件判断目的是只在每个学生开始有成绩时，初始化一次 课程——成绩 字典
            if subject == subjects[0]:
                scores[names[stu]] = {}
            scores[names[stu]][subject] = []
        else:
            stu = random.randrange(studentsNum)
        scores[names[stu]][subject].append(random.randrange(40, 101))

# 每个学生的每门课程的所有成绩记录
recordfn = r'practice\ScoresRecord.xlsx'  # 文件名
wb = Workbook()  # 创建工作簿
ws = wb.active  # 获取自动创建的工作表 sheet
wb.active.title = 'Score List'  # 将自动生成的工作表重命名
ws['A1'] = '姓名'  # 制作表头
ws['B1'] = '课程'
ws['C1'] = '成绩'
row = 2
for name in names:
    ws['A' + str(row)].value = name
    nameStartRow = row
    for subject in subjects:
        ws['B' + str(row)].value = subject
        subjectStartRow = row
        for score in scores[name][subject]:
            ws['C' + str(row)].value = score
            row += 1
        subjectEndRow = row - 1
        ws.merge_cells('B' + str(subjectStartRow) +
                       ':B' + str(subjectEndRow))  # 合并单元格
    nameEndRow = row - 1
    ws.merge_cells('A' + str(nameStartRow) + ':A' + str(nameEndRow))  # 合并单元格
wb.save(recordfn)  # 保存 ScoresRecord.xlsx 文件
wb.close()

# 每个学生的每门课程的最优成绩记录
finalfn = r'practice\Scores.xlsx'  # 文件名
wb = Workbook()  # 创建工作簿
ws = wb.active  # 获取自动创建的工作表 sheet
wb.active.title = 'Score List'  # 将自动生成的工作表重命名
ws['A1'] = '姓名'  # 制作表头
ws['B1'] = '课程'
ws['C1'] = '成绩'
row = 2
for name in names:
    ws['A' + str(row)].value = name
    nameStartRow = row
    for subject in subjects:
        ws['B' + str(row)].value = subject
        subjectStartRow = row
        ws['C' + str(row)].value = max(scores[name][subject])
        row += 1
    nameEndRow = row - 1
    ws.merge_cells('A' + str(nameStartRow) + ':A' + str(nameEndRow))
wb.save(finalfn)  # 保存 Scores.xlsx 文件
wb.close()
