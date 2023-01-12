import pandas as pd
import openpyxl


sheetnames = ['1号线八通线', '2号线', '4号线', '5号线', '6号线', '7号线',
              '8号线', '9号线', '10号线', '11号线', '13号线', '14号线',
              '15号线', '16号线', '17号线', '19号线', '昌平线', '大兴机场线',
              '房山线', '首都机场线', '西郊线', '燕房线', '亦庄线']


# 清洗每个excel文件
def clean_excel(filename):
    # 打开excel文件
    df = pd.read_excel(filename)
    # 给第六列命名
    df.rename(columns={'Unnamed: 5': '房屋户型'}, inplace=True)
    # 获取列名
    columns = df.columns.to_list()
    # 删除 换行 和 空格
    for i in range(len(df)):
        for column in columns:
            df[column][i] = str(df[column][i]).replace(
                '\n', '').replace(' ', '')
    df.to_excel(filename, index=False)


# 为all_info工作簿创建工作表
def creat_new_sheet(filename_write, sheetnames):
    all_info = openpyxl.load_workbook(filename_write)
    if sorted(all_info.sheetnames) == sorted(sheetnames):
        print('工作表均已存在')
        return None
    else:
        print('所需工作表与已有工作表不完全相同，正在重新创建工作表...')
        # 删除工作表
        for sheetname in all_info.sheetnames:
            all_info.remove(all_info[sheetname])
        # 重新创建工作表
        for sheetname in sheetnames:
            all_info.create_sheet(sheetname, 1)
            print('{}工作表创建成功！'.format(sheetname))
    all_info.save(filename_write)


def get_excel_differentsheet(sheetnames, filename_write=r'lianjia\excel\all_info.xlsx'):
    # 防止 df_read.to_excel 覆盖工作表
    writer = pd.ExcelWriter(filename_write)
    # 将各个工作表写在一个工作簿中
    for sheetname in sheetnames:
        df_read = pd.read_excel(r'lianjia\excel\{}.xlsx'.format(sheetname))
        # 将要合并的工作表写入all_info工作簿中
        df_read.to_excel(writer, sheet_name=sheetname, index=False)
        print('{}工作表已写入工作簿'.format(sheetname))
    writer.close()


# 主函数
def main():
    creat_new_sheet(r'lianjia\excel\all_info.xlsx', sheetnames)
    get_excel_differentsheet(sheetnames)


if __name__ == '__main__':
    main()
