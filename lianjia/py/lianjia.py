import requests
import re
from urllib.request import urlopen
from bs4 import BeautifulSoup
import bs4
import openpyxl

a = []
b = []
c = []
d = []
e = []
f = []
g = []
h = []
for i in range(0,1000):
    URL = 'https://bj.lianjia.com/zufang/pg{}'.format(i)
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54'}
    res = requests.get(URL,headers)
    res.encoding = res.apparent_encoding

    soup = bs4.BeautifulSoup(res.text,'lxml')
    # print(soup)
    a = soup.find_all('span',class_="content__list--item-price")
    print(a)

    price=r'<span class="content__list--item-price"><em>(.*?)</em> 元/月</span>'
    res_1=re.findall(price,res.text)
    for i in res_1:
        a.append(i)

    
    quyu=r'<a target="_blank" href=".*?">(.*?)</a>.*?<a title=".*?"'
    res_2=re.findall(quyu,res.text,re.S)
    for i in res_2:
        b.append(i)
# print(b)    
 
    quyu1=r'<a target="_blank" href=".*?">.*?<a href=".*?" target="_blank">(.*?)</a>.*?<a title=".*?"'
    res_3=re.findall(quyu1,res.text,re.S)
    for i in res_3:
        c.append(i)

    louceng=r'<span class="hide">.*?<i>/</i>(.*?)</span>'
    res_4=re.findall(louceng,res.text,re.S)
    for i in res_4:
        d.append(i)

    fangwei=r'<i>/</i>(.*?)\s\s\s\s\s\s\s\s<i>/</i>'
    res_5=re.findall(fangwei,res.text)
    for i in res_5:
        e.append(i)
  
    fangjian=r'\s\s\s\s\s\s\s\s\s\s(.*?)\s\s\s\s\s\s\s\s<span class="hide">'
    res_8=re.findall(fangjian,res.text)
    for i in res_8:
        h.append(i)

    name=r'<a class="twoline" target="_blank" href=".*?">(.*?)</a>'
    res_6=re.findall(name,res.text,re.S)
    for i in res_6:
        f.append(i)   

    mianji=r'(?<=<a target="_blank" href=").*?<i>/</i>(.*?)\s\s\s\s\s\s\s\s<i>/</i>'
    res_7=re.findall(mianji,res.text,re.S)
    for i in res_7:
        g.append(i) 
print(c)
print(b)   
print(h)
print(g)
print(f)   
print(e)
print(d)
print(a)  
with open('text1.xlsx','wb')as fa:
        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet.append(['价格','楼层','方位','名称','面积','','区域','区域1'])
        for i in range(2,len(a)):
                sheet.cell(row=i,column=1,value=a[i])
        for i in range(2,len(d)):
                sheet.cell(row=i,column=2,value=d[i])
        for i in range(2,len(e)):
                sheet.cell(row=i,column=3,value=e[i])
        for i in range(2,len(f)):
                sheet.cell(row=i,column=4,value=f[i])
        for i in range(2,len(g)):
                sheet.cell(row=i,column=5,value=g[i])   
        for i in range(2,len(h)):
                sheet.cell(row=i,column=6,value=h[i])
        for i in range(2,len(b)):
                sheet.cell(row=i,column=7,value=b[i])
        for i in range(2,len(c)):
                sheet.cell(row=i,column=8,value=c[i]) 
        wb.save('text1.xlsx')
