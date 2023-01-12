from openpyxl import Workbook, load_workbook
import requests
import re


def writeInExcel(fn, sheet_name, all_info):
    # 将记录写进excel
    wb = Workbook()  # 创建工作簿
    ws = wb.active  # 获取自动创建的工作表 sheet
    wb.active.title = sheet_name  # 将自动生成的工作表重命名
    ws['A1'] = '每年信息链接'  # 制作表头
    ws['B1'] = '时间'
    ws['C1'] = '均价'

    row = 2
    for info in all_info:
        ws['A' + str(row)].value = info[0]
        ws['B' + str(row)].value = info[1]
        ws['C' + str(row)].value = info[2]
        row += 1
    wb.save(fn)  # 保存文件
    # 按照时间列进行升序排序
    wb.close()


all_info = []
# 模拟浏览器的头部
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54',
    'cookie': 'aQQ_ajkguid=10F11409-85D7-F5F9-8DF2-63E6085E706A; ajk-appVersion=; seo_source_type=1; id58=CrIfoWOnwrwtR+0qeYYaAg==; ctid=14; cmctid=1; wmda_visited_projects=%3B6289197098934; wmda_uuid=09dc92a2d975d280d961bb907da8f084; wmda_new_uuid=1; sessid=2E6F7242-3F42-4339-80C9-406DE68F8355; fzq_h=1c82eae7330be35f2ff97d45c55aa3a1_1672798009960_8dc7b232b74442f4a3411c2a8d1203c9_2029148559; wmda_session_id_6289197098934=1672798013393-1ba4600f-c823-ee1b; obtain_by=2; twe=2; lps=https%3A%2F%2Fbj.zu.anjuke.com%2F%3Ffrom%3DHomePage_TopBar%7Chttps%3A%2F%2Fbeijing.anjuke.com%2F; xxzl_cid=3c9c4d51dd0748df853288f1963d4106; xxzl_deviceid=5q1/DpidANlwOfCzJeIUu5Ip49w+ujVWrTQjWM4rwsf9TFm641cHd1kWBbaAsvcZ'
}
# 由于这个网站要不断访问，所以用 requests 模块来进行爬取
# 如果发现爬取速度不正常，但也没有报错，需要按住 ctrl+单击 链接进行人工验证，再使用爬虫
for year in range(2015, 2023):
    url = 'https://www.anjuke.com/xinfang/fj-bj/' + str(year) + '/'
    # 判断是否爬取成功
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print('爬取失败：{}'.format(response.status_code))
        break
    else:
        print('第' + str(year) + '年爬取成功')

    # 正则表达式匹配
    time_partten = r'<td class="tl"><a href="(.*?)">(.*?)</a></td>'
    price_partten = r'<td>(\d+)元/m²</td>'

    time_info = re.findall(time_partten, response.text, re.S)
    time_link = [i[0] for i in time_info[:12]]
    time = [i[1] for i in time_info[:12]]
    price = re.findall(price_partten, response.text, re.S)

    for i in range(len(time)):
        # 由于爬取的时间格式不统一，需要进行处理
        if time[i][5:] in ['1月', '2月','3月','4月','5月','6月','7月','8月','9月']:
            time[i] = time[i][:5] + '0' + time[i][5:]
        all_info.append([time_link[i],time[i], price[i]])
    
# 将记录写进excel
writeInExcel(r'anjuke_history/history_average.xlsx', '安居客', all_info)
