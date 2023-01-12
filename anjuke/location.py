import requests
import openpyxl

'''
    使用之前，请一定看好注释，还有我的excel结构
'''


def write_title_excel(filename):
    # 写标题
    # 打开excel文件
    wb = openpyxl.load_workbook(filename)
    # 获取工作表，我的工作表下的工作簿名字是 安居客，如果不是，可以自己修改
    ws = wb['安居客']
    if ws['D1'].value == '经度' and ws['E1'].value == '纬度':
        print('标题已经存在')
        pass
    else:
        # 在C列和D列中插入两列
        ws.insert_cols(4, 2)
        # 写入标题
        ws['D1'] = '经度'
        ws['E1'] = '纬度'
    # 保存文件
    wb.save(filename)


# 这个是高德地图的api，有一点不好就是会限制次数，所以我用的是百度地图的api
def get_location_GD(addr):
    para = {'key': '23b10e94bdc3f3fda7e0d0ea57cbf3ab',  # 高德Key，这里你自己申请高德地图api（上网搜一下怎么申请），不要白嫖我的Key
            'address': addr}  # 地址参数
    url = 'https://restapi.amap.com/v3/geocode/geo?'  # 高德地图地理编码API服务地址
    result = requests.get(url, para)  # GET方式请求
    result = result.json()
    # 获取返回参数geocodes中的location，即经纬度
    lon_lat = result['geocodes'][0]['location']
    return lon_lat


# 百度地图api
def get_location_BD(address):
    paramters = {'address': address, 'output': 'json'}
    base = 'http://api.map.baidu.com/geocoder'
    response = requests.get(base, params=paramters)
    answer = response.json()
    # 因为我第一次使用的是高德地图的api，那里的经纬度是一个用逗号隔开的字符串
    # 为了不修改 read_write_location 函数，我把百度地图的api手动拼接成了一个字符串
    return str(answer['result']['location']['lng'])+','+str(answer['result']['location']['lat'])


def read_write_location(filename):
    # 读取地址信息
    # 打开excel文件
    wb = openpyxl.load_workbook(filename)
    # 获取工作表
    ws = wb['安居客']
    # 获取最大行数
    max_row = ws.max_row
    # 读取地址信息
    for i in range(2, max_row + 1):
        # 读取地址信息，并在之前加上 北京市，防止搜索到别的地方相同的小区名称
        # 我的excel文件 B列(第 2 列) 和 C列(第 3 列) 分别是地址和小区名称
        # 所以在读取地址信息的时候，我是把地址和小区名称拼接在一起的
        addr = '北京市' + ws.cell(row=i, column=2).value + ws.cell(row=i, column=3).value
        # lon_lat = get_location_GD(addr)  # 高德地图api，由于限制次数，所以我用的是百度地图api
        lon_lat = get_location_BD(addr)
        # 写入经纬度
        ws.cell(row=i, column=4).value = lon_lat.split(',')[0]
        ws.cell(row=i, column=5).value = lon_lat.split(',')[1]
        print('(' + lon_lat.split(',')[0] + ',' +
              lon_lat.split(',')[1] + ')', end='\t')
        print('第{}条数据写入成功'.format(i-1))
        # 保存文件（牺牲速度，保全次数）
        wb.save(filename)


write_title_excel(r'anjuke\anjuke copy.xlsx')
read_write_location(r'anjuke\anjuke(location)02.xlsx')
