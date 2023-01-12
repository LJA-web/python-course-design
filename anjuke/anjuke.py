from openpyxl import Workbook
import requests
import re


def writeInExcel(fn, sheet_name, all_info):
    # 将记录写进excel
    wb = Workbook()  # 创建工作簿
    ws = wb.active  # 获取自动创建的工作表 sheet
    wb.active.title = sheet_name  # 将自动生成的工作表重命名
    ws['A1'] = '标题'  # 制作表头
    ws['B1'] = '地址'
    ws['C1'] = '小区'
    ws['D1'] = '户型'
    ws['E1'] = '面积'
    ws['F1'] = '楼层'
    ws['G1'] = '月租'
    ws['H1'] = '出租类型'
    ws['I1'] = '朝向'
    ws['J1'] = '有无电梯'
    ws['k1'] = '地铁'
    row = 2
    for info in all_info:
        ws['A' + str(row)].value = info[0]
        ws['B' + str(row)].value = info[1]
        ws['C' + str(row)].value = info[2]
        ws['D' + str(row)].value = info[3]
        ws['E' + str(row)].value = info[4]
        ws['F' + str(row)].value = info[5]
        ws['G' + str(row)].value = info[6]
        ws['H' + str(row)].value = info[7]
        ws['I' + str(row)].value = info[8]
        ws['J' + str(row)].value = info[9]
        ws['K' + str(row)].value = info[10]
        row += 1
    wb.save(fn)  # 保存文件
    wb.close()


all_info = []
# 模拟浏览器的头部
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54',
    'cookie': 'aQQ_ajkguid=10F11409-85D7-F5F9-8DF2-63E6085E706A; ajk-appVersion=; seo_source_type=1; id58=CrIfoWOnwrwtR+0qeYYaAg==; ctid=14; cmctid=1; wmda_visited_projects=%3B6289197098934; wmda_uuid=09dc92a2d975d280d961bb907da8f084; wmda_new_uuid=1; sessid=2E6F7242-3F42-4339-80C9-406DE68F8355; fzq_h=1c82eae7330be35f2ff97d45c55aa3a1_1672798009960_8dc7b232b74442f4a3411c2a8d1203c9_2029148559; wmda_session_id_6289197098934=1672798013393-1ba4600f-c823-ee1b; obtain_by=2; twe=2; lps=https%3A%2F%2Fbj.zu.anjuke.com%2F%3Ffrom%3DHomePage_TopBar%7Chttps%3A%2F%2Fbeijing.anjuke.com%2F; xxzl_cid=3c9c4d51dd0748df853288f1963d4106; xxzl_deviceid=5q1/DpidANlwOfCzJeIUu5Ip49w+ujVWrTQjWM4rwsf9TFm641cHd1kWBbaAsvcZ'
}
# 由于这个网站要不断访问，所以用 requests 模块来进行爬取
# 如果发现爬取速度不正常，但也没有报错，需要按住 ctrl+单击 链接进行人工验证，再使用爬虫
for page in range(1, 51):
    url = 'https://bj.zu.anjuke.com/fangyuan/p' + str(page) + '/'
    # 判断是否爬取成功
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print('爬取失败：{}'.format(response.status_code))
        break
    else:
        print('第' + str(page) + '页爬取成功')
    # try:
    #     response = requests.get(url, headers=headers)
    # except requests.RequestException:
    #     print('爬取失败')
    #     break
    # else:
    #     print('第' + str(page) + '页爬取成功')

    # 正则表达式匹配
    titleAndPrice_pattern = r'<b class="strongbox">(.*?)</b>'
    housetype_pattern = r'<b class="strongbox" style="font-weight: normal;">(.*?)</b>室<b class="strongbox" style="font-weight: normal;">(.*?)</b>厅.*?<i '
    square_pattern = r'<span>\|</span><b class="strongbox" style="font-weight: normal;">(.*?)</b>平米'
    floor_pattern = r'平米<span>\|</span>(.*?)<i '
    location_pattern = r'<a target="_blank" href="https://beijing.anjuke.com/community/view.*?">(.*?)</a>&nbsp;&nbsp;(.*?)</address>'
    tag_pattern = r'<p class="details-item bot-tag".*?>(.*?)</p>'

    titleAndPrice = re.findall(titleAndPrice_pattern, response.text, flags=re.S)
    housetype = re.findall(housetype_pattern, response.text, flags=re.S)
    square = re.findall(square_pattern, response.text, flags=re.S)
    floor = re.findall(floor_pattern, response.text, flags=re.S)
    locations = re.findall(location_pattern, response.text, flags=re.S)
    region, neighbourhood = [], []
    for local in locations:
        region.append(re.sub('\n', '', local[1].strip()))
        neighbourhood.append(local[0])
    # 其他标签，包括出租类型，朝向，有无电梯和地铁
    others = []
    tags = re.findall(tag_pattern, response.text, flags=re.S)
    for tag in tags:
        tag_list = re.findall(r'<span.*?">(.*?)</span>', tag, flags=re.S)
        others.append(tag_list)

    # 将爬取的信息整合
    length = int(len(titleAndPrice)/2)
    for i in range(length):
        info = []
        info.append(titleAndPrice[2*i])
        info.append(region[i])
        info.append(neighbourhood[i])
        info.append(housetype[i][0]+'室'+housetype[i][1]+'厅')
        info.append(square[i]+'平米')
        info.append(floor[i].strip())
        info.append(titleAndPrice[2*i+1])
        tags_length = len(others[i])
        if tags_length == 4:
            info.append(others[i][0])
            info.append(others[i][1])
            info.append(others[i][2][0])
            info.append(others[i][3])
        elif tags_length == 3:
            if '电梯' in others[i][2]:
                info.append(others[i][0])
                info.append(others[i][1])
                info.append('有')
                info.append('无')
            else:
                info.append(others[i][0])
                info.append(others[i][1])
                info.append('无')
                info.append(others[i][2])
        else:
            info.append(others[i][0])
            info.append(others[i][1])
            info.append('无')
            info.append('无')
        all_info.append(info)

# 将信息写入excel
writeInExcel(r'anjuke\anjuke02.xlsx', '安居客', all_info)
